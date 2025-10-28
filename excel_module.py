"""
excel_module.py – Excel integration for ClaimAutomationDragDrop
---------------------------------------------------------------
Handles claim data export, comment annotations, and formatting.
Optimized for reliability, especially with batch OCR runs.
"""

import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
import config

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


def append_to_excel(counselor: str, data: dict, calculations: dict):
    """
    Appends claim data and calculation results to the counselor's Excel sheet.
    
    Args:
        counselor (str): Counselor name (used for filename)
        data (dict): Claim data from OCR extraction
        calculations (dict): Financial calculations from calculations_module
    """
    try:
        # Build path to counselor's Excel file
        excel_path = os.path.join(config.EXCEL_DIR, f"{counselor}.xlsx")
        
        # Load or create workbook
        wb, ws = _load_or_create_excel(excel_path)
        
        # Get next empty row
        next_row = ws.max_row + 1
        
        # Write claim data
        _write_claim_row(ws, next_row, data, calculations)
        
        # Save workbook
        wb.save(excel_path)
        logger.info(f"✅ Excel updated successfully: {os.path.basename(excel_path)}")
        
    except Exception as e:
        logger.exception(f"❌ Excel write error: {e}")
        raise


def _load_or_create_excel(excel_path):
    """Loads an existing workbook or creates a new one."""
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        _write_headers(ws)
    return wb, ws


def _write_headers(ws):
    """Writes column headers if file is new."""
    headers = [
        "Client Name",           # A
        "Client Insurance",      # B
        "Date of Service",       # C
        "Client Copay",          # D - Client payment
        "Deductible being met",  # E - Client payment
        "Insurance Paid",        # F
        "Insurance Contract Amount",  # G = D + E + F
        "65% Counselor Contracted rate", # H = G × 0.65
        "Amount to Counselor (copay and deductible subtracted)", # I = F - (D + E)
        "35% Amount Paid to GWC per claim", # J = G × 0.35
        "Total payout to Counselor",  # K = Running sum of I
        "Remarks"                # L - Remark codes
    ]
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, size=11)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = fill
        cell.font = Font(bold=True, size=11, color="FFFFFF")
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Client Name
    ws.column_dimensions['B'].width = 18  # Insurance
    ws.column_dimensions['C'].width = 15  # Date
    ws.column_dimensions['D'].width = 12  # Copay
    ws.column_dimensions['E'].width = 12  # Deductible
    ws.column_dimensions['F'].width = 14  # Insurance Paid
    ws.column_dimensions['G'].width = 18  # Contract Amount
    ws.column_dimensions['H'].width = 18  # 65% Counselor
    ws.column_dimensions['I'].width = 20  # Amount to Counselor
    ws.column_dimensions['J'].width = 18  # 35% GWC
    ws.column_dimensions['K'].width = 18  # Total Payout
    ws.column_dimensions['L'].width = 40  # Remarks


def _write_claim_row(ws, row, data, calculations):
    """Writes claim data + calculations to Excel."""
    try:
        # Column A: Client Name
        ws.cell(row=row, column=1, value=data.get("Client", "NOTFOUND"))
        
        # Column B: Insurance
        ws.cell(row=row, column=2, value=data.get("Insurance", "NOTFOUND"))
        
        # Column C: Date of Service
        ws.cell(row=row, column=3, value=data.get("Date", "NOTFOUND"))
        
        # Column D: Client Copay (client payment)
        copay_val = _safe_float(data.get("Copay", 0))
        ws.cell(row=row, column=4, value=copay_val)
        
        # Column E: Deductible being met (client payment)
        deductible_val = _safe_float(data.get("Deductible", 0))
        ws.cell(row=row, column=5, value=deductible_val)
        
        # Column F: Insurance Paid
        insurance_payment = _safe_float(data.get("Insurance Payment", 0))
        ws.cell(row=row, column=6, value=insurance_payment)
        
        # Column G: Insurance Contract Amount (D + E + F)
        contracted_rate = calculations.get("contracted_rate", 0)
        ws.cell(row=row, column=7, value=contracted_rate)
        
        # Column H: 65% Keisha Contracted rate
        counselor_65 = calculations.get("counselor_65_percent", 0)
        ws.cell(row=row, column=8, value=counselor_65)
        
        # Column I: Amount to Keisha (copay and deductible subtracted)
        # This is F - (D + E)
        total_payout = calculations.get("total_payout", 0)
        ws.cell(row=row, column=9, value=total_payout)
        
        # Column J: 35% Amount Paid to GWC per claim
        gwc_35 = calculations.get("gwc_35_percent", 0)
        ws.cell(row=row, column=10, value=gwc_35)
        
        # Column K: Total payout to Keisha (running sum)
        # Calculate running sum of column I
        if row == 2:  # First data row
            ws.cell(row=row, column=11, value=total_payout)
        else:
            # Formula: Previous K + Current I
            ws.cell(row=row, column=11, value=f"=K{row-1}+I{row}")
        
        # Column L: Remarks
        remarks = data.get("Remarks", "")
        ws.cell(row=row, column=12, value=remarks)
        
        # Apply formatting and validation comments
        _format_row(ws, row, calculations, data)
        
    except Exception as e:
        logger.error(f"Failed to write row {row}: {e}")
        raise


def _format_row(ws, row, calculations, data):
    """Applies basic formatting and optional validation comments."""
    # Get warnings from calculations and validation
    warnings = []
    
    if calculations.get("warnings"):
        warnings.extend(calculations["warnings"])
    
    if not calculations.get("calculations_valid", True):
        warnings.append("Calculations may be incomplete - check values")
    
    # Check for missing critical data
    if data.get("Client") == "NOTFOUND":
        warnings.append("Client name not found")
    if data.get("Insurance Payment") == "NOTFOUND":
        warnings.append("Insurance payment not found")
    
    # Apply borders
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    
    # Apply formatting to all cells in row
    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        
        # Format currency columns (D, E, F, G, H, I, J, K)
        if col in [4, 5, 6, 7, 8, 9, 10, 11]:
            cell.number_format = '$#,##0.00'
    
    # Add comment if there are warnings
    if warnings:
        comment_text = "\n".join(warnings).strip()
        if comment_text:
            comment = Comment(comment_text, "ClaimAutomation")
            comment.width = 300
            comment.height = 100
            ws.cell(row=row, column=1).comment = comment
            
            # Highlight row with light yellow if there are issues
            yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = yellow_fill


def _safe_float(value) -> float:
    """Safely convert value to float, return 0 if invalid."""
    if value is None or value in ("NOTFOUND", "N/A", ""):
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace("$", "").replace(",", "").strip()
        return float(value)
    except (ValueError, TypeError):
        return 0.0


# ═══════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def get_counselor_excel_path(counselor: str) -> str:
    """Get the full path to a counselor's Excel file."""
    return os.path.join(config.EXCEL_DIR, f"{counselor}.xlsx")


def counselor_file_exists(counselor: str) -> bool:
    """Check if a counselor's Excel file exists."""
    return os.path.exists(get_counselor_excel_path(counselor))


# ═══════════════════════════════════════════════════════════════
# TESTING
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("Testing excel_module.py...")
    
    # Fake test data
    fake_data = {
        "Client": "Jane Doe",
        "Insurance": "Aetna",
        "Date": "10/28/2025",
        "Copay": "25.00",
        "Deductible": "50.00",
        "Insurance Payment": "125.00",
        "Remarks": "PR-3 Copay; PR-1 Deductible"
    }
    
    fake_calc = {
        "contracted_rate": 200.00,
        "counselor_65_percent": 130.00,
        "total_payout": 50.00,  # 125 - (25 + 50)
        "gwc_35_percent": 70.00,
        "calculations_valid": True,
        "warnings": []
    }
    
    # Test writing
    try:
        append_to_excel("TestCounselor", fake_data, fake_calc)
        print("✅ Test Excel written successfully!")
        print(f"   Location: {get_counselor_excel_path('TestCounselor')}")
    except Exception as e:
        print(f"❌ Test failed: {e}")